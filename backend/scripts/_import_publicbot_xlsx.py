"""Import PublicBot-Questions.xlsx into tests/golden/documents/playready_publicbot_v1.json."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "tests" / "golden" / "documents" / "playready_publicbot_v1.json"
OUT_BACKEND = ROOT / "backend" / "tests" / "golden" / "documents" / "playready_publicbot_v1.json"

# Heuristic map: keyword in question → preferred PDF filename(s)
DOC_RULES: list[tuple[list[str], str]] = [
    (["ev certificate", "ocx", "extended validation", "code signing"], "PR_EV_Certificate_Instructions.pdf"),
    (
        [
            "server agreement",
            "server application",
            "server sdk",
            "playready server",
            ".net",
            "licenseservertime",
            "iserverauthorization",
        ],
        "PR_Server_Overview.pdf",
    ),
    (["master agreement", "pima"], "PR_Master_Agreement_Sample_v2013.pdf"),
    (["intermediate product license", "intermediate product", "ipl"], "PR_Intermediate_Product_License_Sample_v2016.pdf"),
    (["final product license", "final product", "fpl"], "PR_Final_Product_License_Sample_v2018.pdf"),
    (["ipla", "licensee portal", "azure b2c", "licensing process", "portal url", "royalty"], "PR_IPLA_Licensing_Portal_FAQ.pdf"),
    (["sl3000", "sl2000", "security level 3000", "security level 2000", "csl", "certificate security level", "tee"], "PR_SL3000_Playbook.pdf"),
    (["robustness", "widely available tools", "device secrets", "pii", "new circumstances", "no circumvention"], "Robustness_Rules_For_PlayReady_Products.pdf"),
    (["compliance", "must understand", "best effort", "output protection", "wm drm", "wmdrm", "root public keys"], "PR_Compliance_Rules_Part01_v2021.pdf"),
    (["live tv", "key rotation", "fifa"], "PR_LiveTV_Protection_Part01_v2015.pdf"),
    (["securestop2", "secure stop", "4.2"], "PR_WhatsNew_v4.2.pdf"),
    (["4.3"], "PR_WhatsNew_v4.3.pdf"),
    (["4.4", "optimized content key", "cdmi"], "PR_WhatsNew_v4.4.pdf"),
    (["4.5", "challenge encryption", "secure time", "key exchange"], "PR_WhatsNew_v4.5.pdf"),
    (["4.6"], "PR_WhatsNew_v4.6.pdf"),
    (["whitepaper", "future shifts", "approved playready"], "PR_Content_Protection_Whitepaper_v2015.pdf"),
    (["device porting kit", "dpk", "ios", "android", "client sdk", "xbox", "silverlight"], "PR_Dev_Clients_Part01_v2015.pdf"),
    (["distribution", "object code"], "PR_Distribution_Overview.pdf"),
    (["metering", "domain join", "domain renew", "revok", "kid", "key id", "encrypt"], "PR_Documentation_Part01.pdf"),
    (["sample agreement"], "Sample Agreements - Microsoft PlayReady.pdf"),
]

STOP = {
    "what",
    "when",
    "where",
    "which",
    "who",
    "how",
    "does",
    "do",
    "did",
    "is",
    "are",
    "the",
    "a",
    "an",
    "to",
    "for",
    "of",
    "in",
    "on",
    "and",
    "or",
    "with",
    "from",
    "your",
    "my",
    "i",
    "me",
    "should",
    "must",
    "can",
    "could",
    "after",
    "before",
    "their",
    "them",
    "this",
    "that",
    "into",
    "about",
    "also",
    "known",
    "as",
    "be",
    "have",
    "has",
    "was",
    "were",
    "will",
    "than",
    "then",
    "between",
    "under",
    "over",
    "per",
    "vs",
    "versus",
}


def map_doc(q: str) -> str | None:
    low = q.lower()
    for keys, doc in DOC_RULES:
        if any(k in low for k in keys):
            return doc
    return None


def category(q: str) -> str:
    low = q.lower()
    if any(k in low for k in ("license", "agreement", "ipla", "portal", "fee", "royalty")):
        return "licensing"
    if any(k in low for k in ("sl3000", "sl2000", "security level", "csl", "tee")):
        return "security_level"
    if "robust" in low or "compliance" in low or "must understand" in low:
        return "compliance"
    if "live tv" in low:
        return "live_tv"
    if re.search(r"4\.\d", low) or "whatsnew" in low or "sdk" in low:
        return "release_notes"
    if any(k in low for k in ("client", "ios", "android", "xbox", "porting")):
        return "client"
    if "server" in low:
        return "server"
    if any(k in low for k in ("encrypt", "kid", "metering", "domain", "revok")):
        return "architecture"
    return "general"


def must_any(q: str) -> list[str]:
    low = q.lower()
    forced: list[str] = []
    for tok in (
        "playready",
        "sl3000",
        "sl2000",
        "securestop2",
        "secure stop",
        "hdcp",
        "ipla",
        "ev certificate",
        "metering",
        "live tv",
        "intermediate product",
        "final product",
        "server agreement",
        "master agreement",
        "robustness",
        "compliance",
        "kid",
        "domain",
        "cbcs",
        "tee",
        "ecp",
    ):
        if tok in low:
            forced.append(tok)
    # Significant tokens from question
    words = re.findall(r"[A-Za-z][A-Za-z0-9\-/]{3,}", q)
    for w in words:
        wl = w.lower()
        if wl in STOP or wl in forced:
            continue
        if wl in {"microsoft", "please", "describe", "explain", "compare", "complete"}:
            continue
        forced.append(wl)
        if len(forced) >= 4:
            break
    # Prefer at least one distinctive term; avoid forcing "playready" alone as only check
    out: list[str] = []
    seen = set()
    for x in forced:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out[:5] or ["playready"]


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
        r"c:\Users\v-pabaheti\Downloads\PublicBot-Questions.xlsx"
    )
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    questions = [
        str(r[0]).strip()
        for r in ws.iter_rows(values_only=True)
        if r and r[0] and str(r[0]).strip() and str(r[0]).strip().lower() != "question"
    ]

    cases = []
    for i, q in enumerate(questions, 1):
        doc = map_doc(q)
        cases.append(
            {
                "id": f"PB{i:03d}",
                "category": category(q),
                "question": q,
                "expect_decision": "answer",
                "expected_answer": (
                    f"Answer from PlayReady Prod Normalised PDFs"
                    + (f" (prefer {doc})" if doc else "")
                    + ". Cross-check in source_document / source_file."
                ),
                "must_any": must_any(q),
                "forbid_any": [],
                "citation_any": [],
                "source_document": doc,
                "source_file": doc,
                "kb_quote_hint": None,
                "map_check": "Imported from PublicBot-Questions.xlsx",
                "origin": "PublicBot-Questions.xlsx",
            }
        )

    # Keep a few refuse traps at end
    for j, (q, expected) in enumerate(
        [
            ("What will PlayReady stock price be next quarter?", "Refuse — not in product docs."),
            ("What is the weather in Redmond tomorrow?", "Refuse — OOD."),
        ],
        1,
    ):
        cases.append(
            {
                "id": f"PB{len(questions) + j:03d}",
                "category": "refuse",
                "question": q,
                "expect_decision": "refuse",
                "expected_answer": expected,
                "must_any": [],
                "forbid_any": [],
                "citation_any": [],
                "source_document": None,
                "source_file": None,
                "kb_quote_hint": None,
                "map_check": "Refuse — OOD",
                "origin": "added_refuse",
            }
        )

    suite = {
        "suite_id": "playready_publicbot_v1",
        "source_kind": "documents",
        "agent_name": "PlayReady Assistant",
        "kb_notes": [
            "Imported from PublicBot-Questions.xlsx (legacy public-bot question set).",
            "source_file is a heuristic PDF mapping into Prod Normalised PDFs — verify in the document.",
            "must_any is derived from the question; tighten after a first eval pass.",
        ],
        "source_xlsx": str(xlsx.name),
        "question_count": len(questions),
        "cases": cases,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(suite, indent=2, ensure_ascii=False), encoding="utf-8")
    OUT_BACKEND.parent.mkdir(parents=True, exist_ok=True)
    OUT_BACKEND.write_text(json.dumps(suite, indent=2, ensure_ascii=False), encoding="utf-8")
    mapped = sum(1 for c in cases if c.get("source_file"))
    print(json.dumps({"questions": len(questions), "cases": len(cases), "mapped_pdf": mapped, "wrote": str(OUT)}, indent=2))


if __name__ == "__main__":
    main()
